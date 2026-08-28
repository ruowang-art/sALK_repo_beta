from __future__ import annotations

import os
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from automouse.exceptions import InventoryUpdateError
from automouse.models import AuditAction, AuditEntry, InventoryUpdateReport, RecordStatus


AUDIT_HEADERS = [
    "Run ID",
    "Timestamp",
    "Sample ID",
    "Mouse ID",
    "Inventory Row",
    "Previous Genotype",
    "Proposed Genotype",
    "Final Genotype",
    "Action",
    "Status",
    "Source File",
    "Source Row",
    "Messages",
]


def _entry_row(entry: AuditEntry) -> list[object]:
    return [
        entry.run_id,
        entry.timestamp,
        entry.sample_id,
        entry.mouse_id or "",
        entry.inventory_row or "",
        entry.previous_genotype or "",
        entry.proposed_genotype or "",
        entry.final_genotype or "",
        entry.action.value,
        entry.status.value,
        entry.source_file,
        entry.source_row,
        " | ".join(entry.messages),
    ]


def _style_data_sheet(sheet: object) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 28
    for index, column in enumerate(sheet.iter_cols(), start=1):
        maximum = max(len(str(cell.value or "")) for cell in list(column)[:250])
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(maximum + 2, 11), 42
        )


def _add_entries_sheet(workbook: Workbook, title: str, entries: Iterable[AuditEntry]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(AUDIT_HEADERS)
    for entry in entries:
        sheet.append(_entry_row(entry))
    _style_data_sheet(sheet)


def write_exception_report(
    report: InventoryUpdateReport,
    output_path: Path,
    *,
    run_id: str,
    raw_record_count: int,
    translated_record_count: int,
    cage_cards_generated: int,
    weaning_groups_generated: int,
) -> Path:
    if output_path.exists():
        raise InventoryUpdateError(f"Refusing to overwrite exception report: {output_path}")
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False
    summary["A1"] = "AutoMouse Run Summary"
    summary["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary.merge_cells("A1:C1")
    summary["A3"] = "Metric"
    summary["B3"] = "Count"
    summary["C3"] = "Notes"
    metrics = [
        ("Run ID", run_id, "Unique execution identifier"),
        ("Raw records", raw_record_count, "Validated Transnetyx rows"),
        ("Translated records", translated_record_count, "Validated R output rows"),
        ("Updated", len(report.updated_mouse_ids), "Blank inventory genotypes filled"),
        ("Confirmed", len(report.confirmed_mouse_ids), "Existing genotype matched"),
        (
            "Mouse not found",
            report.action_counts.get(AuditAction.NOT_FOUND.value, 0),
            "No exact inventory match",
        ),
        (
            "Multiple matches",
            report.action_counts.get(AuditAction.MULTIPLE_MATCHES.value, 0),
            "No automatic update",
        ),
        (
            "Conflicts",
            report.action_counts.get(AuditAction.CONFLICT.value, 0),
            "Existing genotype preserved",
        ),
        (
            "Manual review",
            report.action_counts.get(AuditAction.MANUAL_REVIEW.value, 0),
            "Pending, failed, blank, or invalid",
        ),
        (
            "Card-eligible mice",
            len(report.card_eligible_mouse_ids),
            "Safely matched UPDATED or CONFIRMED mice",
        ),
        (
            "Card grouping issues",
            len(report.card_grouping_exception_mouse_ids),
            "Missing/invalid litter, DOB, or sex metadata",
        ),
        (
            "Litter-sex groups",
            weaning_groups_generated,
            "New-cage groups before five-mouse splitting",
        ),
        ("Cage cards generated", cage_cards_generated, "Live Label rows (max 5 mice)"),
    ]
    for row in metrics:
        summary.append(row)
    for cell in summary[3]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True, color="1F1F1F")
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 28
    summary.column_dimensions["C"].width = 44

    entries = report.audit_entries
    _add_entries_sheet(
        workbook,
        "Mouse Not Found",
        (entry for entry in entries if entry.action == AuditAction.NOT_FOUND),
    )
    _add_entries_sheet(
        workbook,
        "Multiple Matches",
        (entry for entry in entries if entry.action == AuditAction.MULTIPLE_MATCHES),
    )
    _add_entries_sheet(
        workbook,
        "Genotype Conflicts",
        (entry for entry in entries if entry.action == AuditAction.CONFLICT),
    )
    _add_entries_sheet(
        workbook,
        "Duplicate Results",
        (entry for entry in entries if entry.action == AuditAction.DUPLICATE),
    )
    _add_entries_sheet(
        workbook,
        "Pending or Failed",
        (
            entry
            for entry in entries
            if entry.status in {RecordStatus.PENDING_RERUN, RecordStatus.NO_RESULT}
        ),
    )
    _add_entries_sheet(
        workbook,
        "Invalid Genotypes",
        (entry for entry in entries if entry.status == RecordStatus.MANUAL_REVIEW),
    )
    _add_entries_sheet(
        workbook,
        "Card Grouping Issues",
        (
            entry
            for entry in entries
            if (entry.mouse_id or "") in report.card_grouping_exception_mouse_ids
        ),
    )
    _add_entries_sheet(
        workbook,
        "Warnings",
        (entry for entry in entries if entry.messages),
    )
    _add_entries_sheet(workbook, "Audit Log", entries)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_name(f"{output_path.stem}.tmp{output_path.suffix}")
    try:
        workbook.save(temporary)
        reopened = load_workbook(temporary, read_only=True, data_only=False)
        expected_sheets = {
            "Summary",
            "Mouse Not Found",
            "Multiple Matches",
            "Genotype Conflicts",
            "Duplicate Results",
            "Pending or Failed",
            "Invalid Genotypes",
            "Card Grouping Issues",
            "Warnings",
            "Audit Log",
        }
        if set(reopened.sheetnames) != expected_sheets:
            raise InventoryUpdateError(
                "Reopened exception report is missing one or more required sheets."
            )
        reopened.close()
        os.replace(temporary, output_path)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise
    return output_path
