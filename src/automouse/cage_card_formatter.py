from __future__ import annotations

import os
import re
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Border

from automouse.config import CageCardConfig
from automouse.exceptions import CageCardGenerationError, CageCardTemplateError
from automouse.inventory_manager import InventoryTable
from automouse.models import CageCardRecord


def _natural_key(value: str) -> list[Any]:
    return [int(part) if part.isdigit() else part.casefold() for part in re.split(r"(\d+)", value)]


def _parse_date(value: str) -> date | str | None:
    value = value.strip()
    if not value:
        return None
    for date_format in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, date_format).date()
        except ValueError:
            continue
    return value


def _canonical_date(value: str) -> date | None:
    parsed = _parse_date(value)
    return parsed if isinstance(parsed, date) else None


def _normalize_text(value: str) -> str:
    return " ".join(value.strip().split())


def _normalize_sex(value: str) -> str | None:
    normalized = _normalize_text(value).casefold()
    if normalized in {"m", "male"}:
        return "Male"
    if normalized in {"f", "female"}:
        return "Female"
    return None


def _uniform(values: list[str], mixed_value: str = "") -> tuple[str, bool]:
    unique = sorted({value.strip() for value in values if value.strip()})
    if not unique:
        return "", True
    if len(unique) == 1:
        return unique[0], True
    return mixed_value or " / ".join(unique), False


def _parent_genotype(inventory: InventoryTable, parent_value: str) -> str:
    parent_value = parent_value.strip()
    parent_ids = list(dict.fromkeys(re.findall(r"[A-Za-z]+\d+", parent_value)))
    if not parent_ids:
        return ""
    primary_index = inventory.primary_index()
    genotypes: list[str] = []
    for parent_id in parent_ids:
        matches = primary_index.get(parent_id, [])
        if len(matches) != 1:
            return ""
        genotype = inventory.value(matches[0], "genotype")
        if not genotype:
            return ""
        genotypes.append(genotype)
    unique = list(dict.fromkeys(genotypes))
    return unique[0] if len(unique) == 1 else ""


@dataclass(frozen=True, slots=True)
class _CageCandidate:
    row_index: int
    mouse_id: str
    sex: str
    dob: date | None
    strain_key: str
    strain_display: str
    kras_genotype: str


def _strain_value(inventory: InventoryTable, row_index: int) -> str:
    return (
        inventory.value(row_index, "revised_strain")
        or inventory.value(row_index, "strain")
    )


def _strain_key(value: str) -> str:
    return _normalize_text(value).casefold()


def _normalize_kras_genotype(value: str) -> str | None:
    genotype = _normalize_text(value).lstrip("'")
    if not genotype:
        return None
    first_locus = genotype.split(";", 1)[0].strip()
    alleles = [
        allele.strip()
        for allele in first_locus.replace(" ", "").split("/")
        if allele.strip()
    ]
    if len(alleles) != 2:
        return None
    shorthand = {
        "+": "+",
        "K": "K",
        "LSL-G12D": "K",
    }
    normalized = [shorthand.get(allele) for allele in alleles]
    if any(allele is None for allele in normalized):
        return None
    if normalized.count("K") == 1 and normalized.count("+") == 1:
        return "K/+"
    if normalized == ["+", "+"]:
        return "+/+"
    return None


def _dob_window_days(sex: str, config: CageCardConfig) -> int:
    if sex == "Male":
        return config.male_dob_window_days
    if sex == "Female":
        return config.female_dob_window_days
    return 0


def _chunk_rows(rows: list[int], max_mice: int) -> list[list[int]]:
    chunks: list[list[int]] = []
    start = 0
    while start < len(rows):
        remaining = len(rows) - start
        if remaining <= max_mice:
            chunks.append(rows[start:])
            break
        if remaining == max_mice + 1:
            split_size = (remaining + 1) // 2
        elif remaining - max_mice == 1:
            split_size = max_mice - 1
        else:
            split_size = max_mice
        chunks.append(rows[start : start + split_size])
        start += split_size
    return chunks


def _date_range_value(values: list[date]) -> date | str | None:
    unique = sorted(set(values))
    if not unique:
        return None
    if len(unique) == 1:
        return unique[0]
    return f"{unique[0].strftime('%m/%d/%y')} - {unique[-1].strftime('%m/%d/%y')}"


def _offset_date_value(value: date, days: int) -> date:
    return value + timedelta(days=days)


def _date_range_offset(values: list[date], days: int) -> date | str | None:
    unique = sorted(set(values))
    if not unique:
        return None
    if len(unique) == 1:
        return _offset_date_value(unique[0], days)
    start = _offset_date_value(unique[0], days)
    end = _offset_date_value(unique[-1], days)
    return f"{start.strftime('%m/%d/%y')} - {end.strftime('%m/%d/%y')}"


def _derived_wean_date(
    dob_values: list[date],
    wean_date_value: str,
    wean_uniform: bool,
) -> date | str | None:
    unique_dobs = sorted(set(dob_values))
    if len(unique_dobs) > 1:
        return _date_range_offset(unique_dobs, 28)
    if wean_uniform:
        parsed_wean = _parse_date(wean_date_value)
        if isinstance(parsed_wean, date):
            return parsed_wean
    if unique_dobs:
        return _offset_date_value(unique_dobs[0], 28)
    return None


def _litter_key(inventory: InventoryTable, row_index: int) -> tuple[str, str, str] | None:
    mother = _normalize_text(inventory.value(row_index, "mother"))
    father = _normalize_text(inventory.value(row_index, "father"))
    dob = _canonical_date(inventory.value(row_index, "dob"))
    if not mother or not father or dob is None:
        return None
    return (mother, father, dob.isoformat())


def _litter_counts(inventory: InventoryTable) -> dict[tuple[str, str, str], int]:
    counts: dict[tuple[str, str, str], int] = {}
    for row_index, _ in enumerate(inventory.rows):
        key = _litter_key(inventory, row_index)
        if key:
            counts[key] = counts.get(key, 0) + 1
    return counts


def _build_litter_sex_records(
    inventory: InventoryTable,
    eligible_mouse_ids: set[str],
    config: CageCardConfig,
) -> tuple[list[CageCardRecord], dict[str, str], list[str]]:
    """Build new weaning-cage rows from litter, sex, and a five-mouse limit."""
    primary_index = inventory.primary_index()
    grouping_issues: dict[str, str] = {}
    warnings: list[str] = []
    grouped_rows: dict[tuple[str, str, str, str], list[int]] = {}

    for mouse_id in sorted(eligible_mouse_ids, key=_natural_key):
        matches = primary_index.get(mouse_id, [])
        if len(matches) != 1:
            grouping_issues[mouse_id] = (
                "Mouse does not have exactly one primary inventory row."
            )
            continue
        row_index = matches[0]
        mother = _normalize_text(inventory.value(row_index, "mother"))
        father = _normalize_text(inventory.value(row_index, "father"))
        dob_value = inventory.value(row_index, "dob")
        dob = _canonical_date(dob_value)
        sex_value = inventory.value(row_index, "sex")
        sex = _normalize_sex(sex_value)
        required_values = {
            "Mother": mother,
            "Father": father,
            "DOB": dob_value,
            "Sex": sex_value,
        }
        missing_fields = [name for name, value in required_values.items() if not value]
        if dob_value and dob is None:
            missing_fields.append("valid DOB")
        if sex_value and sex is None:
            missing_fields.append("recognized Sex")
        if missing_fields:
            grouping_issues[mouse_id] = (
                "Missing required weaning-card field(s): "
                + ", ".join(missing_fields)
            )
            continue
        key = (mother, father, dob.isoformat(), sex)
        grouped_rows.setdefault(key, []).append(row_index)

    litter_counts = _litter_counts(inventory)

    records: list[CageCardRecord] = []
    for group_key in sorted(grouped_rows, key=lambda key: _natural_key("|".join(key))):
        mother, father, dob_key, normalized_sex = group_key
        rows = sorted(
            grouped_rows[group_key],
            key=lambda index: _natural_key(inventory.value(index, "mouse_id")),
        )
        litter_key = (mother, father, dob_key)
        litter_count = litter_counts.get(litter_key, len(rows))
        for start in range(0, len(rows), config.max_mice_per_card):
            chunk = rows[start : start + config.max_mice_per_card]
            strains = [_strain_value(inventory, index) for index in chunk]
            strain, strain_uniform = _uniform(strains)
            sexes = [inventory.value(index, "sex") for index in chunk]
            sex = normalized_sex
            sex_uniform = all(_normalize_sex(value) == normalized_sex for value in sexes)
            mothers = [inventory.value(index, "mother") for index in chunk]
            mother, mother_uniform = _uniform(mothers)
            fathers = [inventory.value(index, "father") for index in chunk]
            father, father_uniform = _uniform(fathers)
            dobs = [_canonical_date(inventory.value(index, "dob")) for index in chunk]
            dob_uniform = all(
                value is not None and value.isoformat() == dob_key for value in dobs
            )
            wean_dates = [inventory.value(index, "wean_date") for index in chunk]
            wean_date, wean_uniform = _uniform(wean_dates)
            row_warnings: list[str] = []
            for label, uniform in (
                ("strain", strain_uniform),
                ("sex", sex_uniform),
                ("mother", mother_uniform),
                ("father", father_uniform),
                ("date of birth", dob_uniform),
                ("wean date", wean_uniform),
            ):
                if not uniform:
                    row_warnings.append(
                        f"Mixed {label} values in weaning group {' | '.join(group_key)}."
                    )
            warnings.extend(row_warnings)
            records.append(
                CageCardRecord(
                    cage_id="|".join(group_key),
                    experiment_url=config.experiment_url,
                    strain=strain,
                    animal_count=len(chunk),
                    litter_count=litter_count,
                    sex=sex,
                    date_weaned=_derived_wean_date(dobs, wean_date, wean_uniform),
                    date_born=date.fromisoformat(dob_key),
                    mouse_ids=[inventory.value(index, "mouse_id") for index in chunk],
                    genotypes=[inventory.value(index, "genotype") for index in chunk],
                    dam=mother if mother_uniform else "",
                    dam_genotype=_parent_genotype(inventory, mother) if mother_uniform else "",
                    sire=father if father_uniform else "",
                    sire_genotype=_parent_genotype(inventory, father) if father_uniform else "",
                    warnings=row_warnings,
                )
            )
    for mouse_id, message in grouping_issues.items():
        warnings.append(f"{mouse_id}: {message}")
    records.sort(
        key=lambda record: _natural_key(record.mouse_ids[0] if record.mouse_ids else "")
    )
    return records, grouping_issues, sorted(set(warnings))


def _build_compatible_cage_records(
    inventory: InventoryTable,
    eligible_mouse_ids: set[str],
    config: CageCardConfig,
) -> tuple[list[CageCardRecord], dict[str, str], list[str]]:
    """Build new cage rows by sex, strain, Kras genotype, and DOB window."""
    primary_index = inventory.primary_index()
    grouping_issues: dict[str, str] = {}
    warnings: list[str] = []
    grouped_candidates: dict[tuple[str, str, str], list[_CageCandidate]] = {}

    for mouse_id in sorted(eligible_mouse_ids, key=_natural_key):
        matches = primary_index.get(mouse_id, [])
        if len(matches) != 1:
            grouping_issues[mouse_id] = (
                "Mouse does not have exactly one primary inventory row."
            )
            continue
        row_index = matches[0]
        dob_value = inventory.value(row_index, "dob")
        dob = _canonical_date(dob_value)
        sex_value = inventory.value(row_index, "sex")
        sex = _normalize_sex(sex_value)
        strain_display = _normalize_text(_strain_value(inventory, row_index))
        kras_genotype = _normalize_kras_genotype(inventory.value(row_index, "genotype"))
        required_values = {
            "Sex": sex_value,
            "Strain": strain_display,
            "Genotype": inventory.value(row_index, "genotype"),
        }
        missing_fields = [name for name, value in required_values.items() if not value]
        if dob_value and dob is None:
            missing_fields.append("valid DOB")
        if sex_value and sex is None:
            missing_fields.append("recognized Sex")
        if required_values["Genotype"] and kras_genotype is None:
            missing_fields.append("recognized Kras genotype (K/+ or +/+)")
        if missing_fields:
            grouping_issues[mouse_id] = (
                "Missing required compatible-cage field(s): "
                + ", ".join(missing_fields)
            )
            continue
        assert sex is not None
        assert kras_genotype is not None
        candidate = _CageCandidate(
            row_index=row_index,
            mouse_id=mouse_id,
            sex=sex,
            dob=dob,
            strain_key=_strain_key(strain_display),
            strain_display=strain_display,
            kras_genotype=kras_genotype,
        )
        key = (candidate.sex, candidate.strain_key, candidate.kras_genotype)
        grouped_candidates.setdefault(key, []).append(candidate)

    litter_counts = _litter_counts(inventory)
    records: list[CageCardRecord] = []
    for group_key in sorted(grouped_candidates, key=lambda key: _natural_key("|".join(key))):
        known_dob_candidates = [
            candidate for candidate in grouped_candidates[group_key] if candidate.dob is not None
        ]
        unknown_dob_candidates = [
            candidate for candidate in grouped_candidates[group_key] if candidate.dob is None
        ]
        candidates = sorted(
            known_dob_candidates,
            key=lambda candidate: (candidate.dob or date.max, _natural_key(candidate.mouse_id)),
        )
        clusters: list[list[_CageCandidate]] = []
        current: list[_CageCandidate] = []
        for candidate in candidates:
            if not current:
                current = [candidate]
                continue
            window_days = _dob_window_days(candidate.sex, config)
            if (candidate.dob - current[0].dob).days <= window_days:
                current.append(candidate)
            else:
                clusters.append(current)
                current = [candidate]
        if current:
            clusters.append(current)
        if unknown_dob_candidates:
            clusters.append(
                sorted(
                    unknown_dob_candidates,
                    key=lambda candidate: _natural_key(candidate.mouse_id),
                )
            )

        for cluster in clusters:
            row_indexes = [candidate.row_index for candidate in cluster]
            sorted_rows = sorted(
                row_indexes,
                key=lambda index: (
                    _canonical_date(inventory.value(index, "dob")) or date.min,
                    _natural_key(inventory.value(index, "mouse_id")),
                ),
            )
            for chunk in _chunk_rows(sorted_rows, config.max_mice_per_card):
                strains = [_strain_value(inventory, index) for index in chunk]
                strain_keys = [_strain_key(value) for value in strains]
                strain = _normalize_text(strains[0]) if strains else ""
                strain_uniform = len(set(strain_keys)) <= 1
                sexes = [inventory.value(index, "sex") for index in chunk]
                normalized_sexes = [_normalize_sex(value) for value in sexes]
                sex_uniform = len(set(normalized_sexes)) == 1
                sex = normalized_sexes[0] if normalized_sexes and normalized_sexes[0] else ""
                kras_values = [
                    _normalize_kras_genotype(inventory.value(index, "genotype"))
                    for index in chunk
                ]
                kras_uniform = len(set(kras_values)) == 1
                mothers = [inventory.value(index, "mother") for index in chunk]
                mother, mother_uniform = _uniform(mothers)
                fathers = [inventory.value(index, "father") for index in chunk]
                father, father_uniform = _uniform(fathers)
                dobs = [
                    dob
                    for dob in (
                        _canonical_date(inventory.value(index, "dob"))
                        for index in chunk
                    )
                    if dob is not None
                ]
                date_born = _date_range_value(dobs)
                dob_span = (max(dobs) - min(dobs)).days if dobs else 0
                dob_uniform = dob_span <= _dob_window_days(sex, config)
                wean_dates = [inventory.value(index, "wean_date") for index in chunk]
                wean_date, wean_uniform = _uniform(wean_dates)
                represented_litters = {
                    key for key in (_litter_key(inventory, index) for index in chunk) if key
                }
                litter_count = sum(litter_counts.get(key, 0) for key in represented_litters)
                if not litter_count:
                    litter_count = len(chunk)

                row_warnings: list[str] = []
                base_cage_label = (
                    f"{sex}|{strain}|{kras_values[0] or ''}|"
                    f"{date_born if date_born else ''}"
                )
                if not dobs:
                    row_warnings.append(
                        f"Compatible cage {base_cage_label} has no DOB metadata; "
                        "DATE BORN and DATE WEANED are blank."
                    )
                mouse_values = [inventory.value(index, "mouse_id") for index in chunk]
                cage_id = f"{base_cage_label}|{'-'.join(mouse_values)}"
                for label, uniform in (
                    ("strain", strain_uniform),
                    ("sex", sex_uniform),
                    ("Kras genotype", kras_uniform),
                    ("date of birth window", dob_uniform),
                    ("mother", mother_uniform),
                    ("father", father_uniform),
                    ("wean date", wean_uniform),
                ):
                    if not uniform:
                        row_warnings.append(
                            f"Mixed {label} values in compatible cage {base_cage_label}."
                        )
                if len(represented_litters) > 1:
                    row_warnings.append(
                        f"Compatible cage {base_cage_label} contains "
                        f"{len(represented_litters)} source litters."
                    )
                warnings.extend(row_warnings)
                records.append(
                    CageCardRecord(
                        cage_id=cage_id,
                        experiment_url=config.experiment_url,
                        strain=strain,
                        animal_count=len(chunk),
                        litter_count=litter_count,
                        sex=sex,
                        date_weaned=_derived_wean_date(dobs, wean_date, wean_uniform),
                        date_born=date_born,
                        mouse_ids=mouse_values,
                        genotypes=[inventory.value(index, "genotype") for index in chunk],
                        dam=mother if mother_uniform else "",
                        dam_genotype=_parent_genotype(inventory, mother) if mother_uniform else "",
                        sire=father if father_uniform else "",
                        sire_genotype=_parent_genotype(inventory, father) if father_uniform else "",
                        warnings=row_warnings,
                    )
                )
    for mouse_id, message in grouping_issues.items():
        warnings.append(f"{mouse_id}: {message}")
    records.sort(
        key=lambda record: _natural_key(record.mouse_ids[0] if record.mouse_ids else "")
    )
    return records, grouping_issues, sorted(set(warnings))


def build_cage_card_records(
    inventory: InventoryTable,
    eligible_mouse_ids: set[str],
    config: CageCardConfig,
) -> tuple[list[CageCardRecord], dict[str, str], list[str]]:
    if config.grouping_strategy == "weaning_litter_sex":
        return _build_litter_sex_records(inventory, eligible_mouse_ids, config)
    return _build_compatible_cage_records(inventory, eligible_mouse_ids, config)


def generate_cage_cards(
    records: list[CageCardRecord],
    template_path: Path,
    output_path: Path,
    config: CageCardConfig,
) -> Path:
    if not template_path.is_file():
        raise CageCardTemplateError(f"Cage-card template not found: {template_path}")
    if output_path.exists():
        raise CageCardGenerationError(f"Refusing to overwrite cage-card output: {output_path}")

    workbook = load_workbook(template_path)
    if config.sheet_name not in workbook.sheetnames:
        raise CageCardTemplateError(
            f"Template sheet {config.sheet_name!r} not found in {template_path}."
        )
    sheet = workbook[config.sheet_name]
    actual_headers = [sheet.cell(1, column).value or "" for column in range(1, 24)]
    if config.expected_headers and tuple(actual_headers) != config.expected_headers:
        raise CageCardTemplateError(
            "Cage-card template headers do not match configuration. "
            f"Actual: {actual_headers}"
        )

    original_max_row = sheet.max_row
    for row in range(2, max(original_max_row, len(records) + 1) + 1):
        for column in range(1, 24):
            sheet.cell(row, column).value = None

    for record_index, record in enumerate(records, start=2):
        if record_index > sheet.max_row:
            source_row = 2
            for column in range(1, 24):
                source = sheet.cell(source_row, column)
                target = sheet.cell(record_index, column)
                if source.has_style:
                    target._style = copy(source._style)
                target.number_format = source.number_format
                target.alignment = copy(source.alignment)
                target.border = copy(source.border)
                target.fill = copy(source.fill)
                target.font = copy(source.font)
        for column, value in enumerate(
            record.template_row(config.max_mice_per_card), start=1
        ):
            sheet.cell(record_index, column).value = value

    # Strip any leftover cell borders (e.g. stray formatting artifacts in the
    # template) from rows below the last populated card row, so blank rows
    # in the output never show a border.
    last_populated_row = len(records) + 1
    for row in range(last_populated_row + 1, max(original_max_row, last_populated_row) + 1):
        for column in range(1, 24):
            sheet.cell(row, column).border = Border()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_name(f"{output_path.stem}.tmp{output_path.suffix}")
    try:
        workbook.save(temporary)
        reopened = load_workbook(temporary, read_only=True, data_only=False)
        reopened_sheet = reopened[config.sheet_name]
        reopened_headers = [
            reopened_sheet.cell(1, column).value or "" for column in range(1, 24)
        ]
        if reopened_headers != actual_headers:
            raise CageCardGenerationError("Cage-card headers changed during save.")
        populated = sum(
            any(reopened_sheet.cell(row, column).value not in (None, "") for column in range(1, 24))
            for row in range(2, len(records) + 2)
        )
        if populated != len(records):
            raise CageCardGenerationError(
                f"Expected {len(records)} populated cage-card row(s), found {populated}."
            )
        reopened.close()
        os.replace(temporary, output_path)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise
    return output_path
