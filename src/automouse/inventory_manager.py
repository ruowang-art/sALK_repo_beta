from __future__ import annotations

import csv
import os
import re
import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from automouse.config import InventoryConfig
from automouse.exceptions import InventoryUpdateError, InventoryValidationError
from automouse.input_manager import calculate_sha256
from automouse.models import AuditAction, AuditEntry, RecordStatus


@dataclass(slots=True)
class InventoryTable:
    source_path: Path
    headers: list[str]
    rows: list[list[str]]
    config: InventoryConfig

    def value(self, row_index: int, role: str) -> str:
        return self.rows[row_index][self.config.column_index(role)].strip()

    def set_value(self, row_index: int, role: str, value: str) -> None:
        self.rows[row_index][self.config.column_index(role)] = value

    def identifier_index(self) -> dict[str, list[int]]:
        index: dict[str, list[int]] = {}
        for row_index, _ in enumerate(self.rows):
            identifiers = {
                self.value(row_index, role)
                for role in self.config.identifier_roles
                if role in self.config.columns
            }
            for identifier in identifiers:
                if identifier:
                    index.setdefault(identifier, []).append(row_index)
        return index

    def primary_index(self) -> dict[str, list[int]]:
        index: dict[str, list[int]] = {}
        for row_index, _ in enumerate(self.rows):
            identifier = self.value(row_index, "mouse_id")
            if identifier:
                index.setdefault(identifier, []).append(row_index)
        return index


def load_inventory(path: Path, config: InventoryConfig) -> InventoryTable:
    if not path.is_file():
        raise InventoryValidationError(f"Inventory file does not exist: {path}")
    if config.format != "csv":
        raise InventoryValidationError(
            f"Unsupported inventory format {config.format!r}; expected 'csv'."
        )

    try:
        with path.open(encoding="utf-8-sig", newline="") as stream:
            reader = csv.reader(stream, strict=True)
            headers = next(reader)
            raw_rows = list(reader)
    except StopIteration as error:
        raise InventoryValidationError(f"Inventory has no header row: {path}") from error
    except (OSError, UnicodeDecodeError, csv.Error) as error:
        raise InventoryValidationError(f"Unable to read inventory {path}: {error}") from error

    errors: list[str] = []
    maximum_position = max(config.columns.values(), default=0)
    if maximum_position > len(headers):
        errors.append(
            f"Configured inventory column {maximum_position} exceeds the "
            f"{len(headers)} available columns."
        )
    for role, expected_header in config.expected_headers.items():
        if role not in config.columns:
            errors.append(f"Expected-header role is not in inventory.columns: {role}")
            continue
        index = config.column_index(role)
        if index >= len(headers):
            continue
        actual = headers[index].strip()
        if actual.casefold() != expected_header.strip().casefold():
            errors.append(
                f"Inventory column {index + 1} for {role!r} is {actual!r}; "
                f"expected {expected_header!r}."
            )

    width = len(headers)
    rows: list[list[str]] = []
    for row_number, row in enumerate(raw_rows, start=2):
        if len(row) > width:
            errors.append(
                f"Inventory row {row_number} has {len(row)} fields; header has {width}."
            )
            continue
        rows.append(row + [""] * (width - len(row)))

    if errors:
        raise InventoryValidationError(
            "Inventory validation failed:\n- " + "\n- ".join(errors)
        )

    table = InventoryTable(path, headers, rows, config)
    duplicate_primary = sorted(
        identifier
        for identifier, matches in table.primary_index().items()
        if len(matches) > 1
    )
    if duplicate_primary:
        raise InventoryValidationError(
            "Duplicate primary mouse identifiers in inventory: "
            + ", ".join(duplicate_primary[:20])
        )
    return table


def backup_inventory(inventory_path: Path, backup_path: Path) -> Path:
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    if backup_path.exists():
        raise InventoryUpdateError(f"Refusing to overwrite inventory backup: {backup_path}")
    shutil.copy2(inventory_path, backup_path)
    if calculate_sha256(inventory_path) != calculate_sha256(backup_path):
        backup_path.unlink(missing_ok=True)
        raise InventoryUpdateError(
            f"Inventory backup checksum does not match source: {backup_path}"
        )
    return backup_path


def ensure_audit_columns(table: InventoryTable) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for role, header in table.config.audit_column_names.items():
        if header in table.headers:
            index = table.headers.index(header)
        else:
            index = len(table.headers)
            table.headers.append(header)
            for row in table.rows:
                row.append("")
        indexes[role] = index
    return indexes


def _natural_identifier_key(value: str) -> tuple[int, list[tuple[int, int | str]]]:
    value = value.strip()
    if not value:
        return (1, [])
    parts: list[tuple[int, int | str]] = [
        (0, int(part)) if part.isdigit() else (1, part.casefold())
        for part in re.split(r"(\d+)", value)
        if part
    ]
    return (0, parts)


def _inventory_row_sort_key(
    table: InventoryTable, row: list[str]
) -> tuple[int, list[tuple[int, int | str]]]:
    index = table.config.column_index("mouse_id")
    value = row[index] if index < len(row) else ""
    return _natural_identifier_key(value)


def save_updated_inventory(
    table: InventoryTable,
    csv_output_path: Path,
    xlsx_output_path: Path,
    *,
    updated_values: dict[str, str] | None = None,
) -> tuple[Path, Path]:
    csv_output_path.parent.mkdir(parents=True, exist_ok=True)
    xlsx_output_path.parent.mkdir(parents=True, exist_ok=True)
    if csv_output_path.exists() or xlsx_output_path.exists():
        raise InventoryUpdateError("Refusing to overwrite an existing inventory output.")

    output_rows = sorted(table.rows, key=lambda row: _inventory_row_sort_key(table, row))
    csv_temporary = csv_output_path.with_suffix(csv_output_path.suffix + ".tmp")
    try:
        with csv_temporary.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.writer(stream, lineterminator="\n")
            writer.writerow(table.headers)
            writer.writerows(output_rows)
        with csv_temporary.open(encoding="utf-8-sig", newline="") as stream:
            reader = csv.reader(stream, strict=True)
            saved_headers = next(reader)
            saved_rows = list(reader)
        if saved_headers != table.headers or len(saved_rows) != len(output_rows):
            raise InventoryUpdateError(
                "Reopened inventory CSV does not match the in-memory row/header counts."
            )
        os.replace(csv_temporary, csv_output_path)
    except Exception:
        csv_temporary.unlink(missing_ok=True)
        raise

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = table.config.output_sheet_name
    sheet.append(table.headers)
    for row in output_rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(table.headers))}{len(output_rows) + 1}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 28

    sample_rows = output_rows[: min(len(output_rows), 1500)]
    for column_index, header in enumerate(table.headers, start=1):
        maximum = max(
            [len(str(header))]
            + [len(str(row[column_index - 1])) for row in sample_rows]
        )
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(maximum + 2, 10), 36
        )
    for role, width in {
        "last_updated": 30,
        "genotype_source": 56,
        "run_id": 26,
    }.items():
        header = table.config.audit_column_names.get(role)
        if header and header in table.headers:
            sheet.column_dimensions[
                get_column_letter(table.headers.index(header) + 1)
            ].width = width

    xlsx_temporary = xlsx_output_path.with_name(
        f"{xlsx_output_path.stem}.tmp{xlsx_output_path.suffix}"
    )
    try:
        workbook.save(xlsx_temporary)
        reopened = load_workbook(xlsx_temporary, read_only=True, data_only=False)
        reopened_sheet = reopened[table.config.output_sheet_name]
        if reopened_sheet.max_row != len(output_rows) + 1:
            raise InventoryUpdateError(
                "Reopened inventory workbook has an unexpected row count."
            )
        if reopened_sheet.max_column != len(table.headers):
            raise InventoryUpdateError(
                "Reopened inventory workbook has an unexpected column count."
            )
        if updated_values:
            mouse_col = table.config.column_index("mouse_id") + 1
            genotype_col = table.config.column_index("genotype") + 1
            seen: dict[str, str] = {}
            for row in reopened_sheet.iter_rows(
                min_row=2,
                min_col=min(mouse_col, genotype_col),
                max_col=max(mouse_col, genotype_col),
                values_only=True,
            ):
                if mouse_col < genotype_col:
                    mouse_id, genotype = row[0], row[-1]
                else:
                    genotype, mouse_id = row[0], row[-1]
                if mouse_id in updated_values:
                    seen[str(mouse_id)] = "" if genotype is None else str(genotype)
            missing = sorted(set(updated_values) - set(seen))
            incorrect = sorted(
                key for key, expected in updated_values.items() if seen.get(key) != expected
            )
            if missing or incorrect:
                raise InventoryUpdateError(
                    f"Updated inventory verification failed; missing={missing}, "
                    f"incorrect={incorrect}."
                )
        reopened.close()
        os.replace(xlsx_temporary, xlsx_output_path)
    except Exception:
        xlsx_temporary.unlink(missing_ok=True)
        raise
    return csv_output_path, xlsx_output_path


def write_audit_csv(entries: list[AuditEntry], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        raise InventoryUpdateError(f"Refusing to overwrite audit file: {output_path}")
    headers = [
        "run_id",
        "timestamp",
        "sample_id",
        "mouse_id",
        "inventory_row",
        "previous_genotype",
        "proposed_genotype",
        "final_genotype",
        "action",
        "status",
        "source_file",
        "source_row",
        "messages",
    ]
    temporary = output_path.with_suffix(output_path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=headers, lineterminator="\n")
        writer.writeheader()
        writer.writerows(entry.to_dict() for entry in entries)
    os.replace(temporary, output_path)
    return output_path


def read_audit_csv(path: Path) -> list[AuditEntry]:
    """Reconstruct the audit trail a prior stage wrote with ``write_audit_csv``.

    This is what lets ``generate-cards`` run as its own, later, independent
    step: everything ``write_exception_report`` needs about a run's records
    (action/status per mouse) is recoverable from this file alone, so a
    separate cage-card stage never needs to re-run matching to know what an
    earlier inventory-update stage decided.
    """
    if not path.is_file():
        raise InventoryUpdateError(f"Audit file not found: {path}")
    entries: list[AuditEntry] = []
    with path.open(newline="", encoding="utf-8") as stream:
        for row in csv.DictReader(stream):
            entries.append(
                AuditEntry(
                    run_id=row["run_id"],
                    timestamp=row["timestamp"],
                    sample_id=row["sample_id"],
                    mouse_id=row["mouse_id"] or None,
                    inventory_row=int(row["inventory_row"]) if row["inventory_row"] else None,
                    previous_genotype=row["previous_genotype"] or None,
                    proposed_genotype=row["proposed_genotype"] or None,
                    final_genotype=row["final_genotype"] or None,
                    action=AuditAction(row["action"]),
                    status=RecordStatus(row["status"]),
                    source_file=row["source_file"],
                    source_row=int(row["source_row"]) if row["source_row"] else 0,
                    messages=row["messages"].split(" | ") if row["messages"] else [],
                )
            )
    return entries
