"""Reads Live Label cage-card workbooks (as produced by Möuseley Kräs) into
flat per-mouse records, ready for cross-workbook consolidation.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from xolpotsxol.exceptions import CageCardFormatError
from xolpotsxol.models import EXPECTED_HEADERS, SourceMouse

_DATE_FORMATS = ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d")


def _parse_single_date(value: str) -> date | None:
    value = value.strip()
    if not value:
        return None
    for date_format in _DATE_FORMATS:
        try:
            return datetime.strptime(value, date_format).date()
        except ValueError:
            continue
    return None


def _parse_dob_cell(value: object) -> tuple[date | None, date | None]:
    if isinstance(value, datetime):
        return value.date(), value.date()
    if isinstance(value, date):
        return value, value
    text = str(value or "").strip()
    if not text:
        return None, None
    if " - " in text:
        start_text, end_text = text.split(" - ", 1)
        start = _parse_single_date(start_text)
        end = _parse_single_date(end_text)
        if start is not None and end is not None:
            return start, end
        return None, None
    parsed = _parse_single_date(text)
    return parsed, parsed


def read_cage_card_workbook(path: Path, filename: str) -> tuple[list[SourceMouse], list[str]]:
    """Parse one uploaded Live Label workbook.

    Returns ``(mice, warnings)``. Raises :class:`CageCardFormatError` if the
    workbook's header row does not match the expected Live Label layout —
    Xol-Pots-Xol never guesses at a different column layout.
    """
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    header = tuple(str(sheet.cell(1, column).value or "").strip() for column in range(1, 24))
    if header != EXPECTED_HEADERS:
        raise CageCardFormatError(
            f"{filename}: this does not look like a Live Label cage-card workbook "
            "produced by Möuseley Kräs (the header row does not match)."
        )

    mice: list[SourceMouse] = []
    warnings: list[str] = []
    for row_index in range(2, sheet.max_row + 1):
        values = [sheet.cell(row_index, column).value for column in range(1, 24)]
        if all(value in (None, "") for value in values):
            continue

        experiment_url = str(values[0] or "").strip()
        strain = str(values[1] or "").strip()
        sex = str(values[4] or "").strip()
        dob_min, dob_max = _parse_dob_cell(values[6])
        try:
            in_litter = int(values[3]) if values[3] not in (None, "") else 1
        except (TypeError, ValueError):
            in_litter = 1
        dam = str(values[17] or "").strip()
        dam_genotype = str(values[18] or "").strip()
        sire = str(values[19] or "").strip()
        sire_genotype = str(values[20] or "").strip()
        breeder = str(values[21] or "").strip()

        row_mice: list[SourceMouse] = []
        for slot in range(5):
            mouse_id = str(values[7 + slot] or "").strip()
            if not mouse_id:
                continue
            genotype = str(values[12 + slot] or "").strip()
            row_mice.append(
                SourceMouse(
                    mouse_id=mouse_id,
                    genotype=genotype,
                    sex=sex,
                    strain=strain,
                    dob_min=dob_min,
                    dob_max=dob_max,
                    dam=dam,
                    dam_genotype=dam_genotype,
                    sire=sire,
                    sire_genotype=sire_genotype,
                    breeder=breeder,
                    experiment_url=experiment_url,
                    source_file=filename,
                    source_row=row_index,
                    source_in_litter=in_litter,
                )
            )
        if not row_mice:
            warnings.append(f"{filename} row {row_index}: no mice found in MOUSE 1-5; skipped.")
            continue
        if not sex or not strain:
            warnings.append(
                f"{filename} row {row_index}: missing sex/strain; its mice were kept "
                "unconsolidated in the output."
            )
        mice.extend(row_mice)
    return mice, warnings
