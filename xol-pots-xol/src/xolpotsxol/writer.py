"""Writes a new Live Label-format cage-card workbook from a
:class:`~xolpotsxol.models.ConsolidationResult`.

This never touches the actual Möuseley Kräs cage-card template file — the
header layout is a small, fixed constant (:data:`EXPECTED_HEADERS`), and the
output workbook is built from scratch with openpyxl.
"""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from xolpotsxol.consolidator import KRAS_GENOTYPE_GRAMMAR_VERSION, unconsolidated_reasons
from xolpotsxol.models import EXPECTED_HEADERS, MAX_MICE_PER_CAGE, ConsolidationResult, SourceMouse

_REVIEW_HEADERS = (
    "Source File", "Source Row", "Mouse ID", "Strain", "Sex",
    "Raw Genotype", "Reason(s)",
)

# A freshly created openpyxl sheet defaults every column to ~8.43 characters
# wide, which is too narrow for a date (e.g. "2026-01-19"), a DOB/wean-date
# range (e.g. "01/19/26 - 01/21/26"), or a composite genotype string (e.g.
# "+/+; L/L; T/T; HC9/HC9") — Excel renders an overflowing numeric/date cell
# as a column of "#" characters rather than truncating it, which is exactly
# the "strings of #'s" symptom this fixes. The real Möuseley Kräs cage-card
# template already has wide-enough columns baked in; this workbook is built
# from scratch, so a fixed guess per column can still fall short once real
# data (e.g. a wider DOB range) is written — instead, widths are computed
# from the actual longest value in each column after every row is written.
_MIN_COLUMN_WIDTH = 10
_COLUMN_WIDTH_PADDING = 3


def _display_length(value: object) -> int:
    if value is None or value == "":
        return 0
    if hasattr(value, "strftime"):
        return len(value.strftime("%Y-%m-%d"))
    return len(str(value))


def _autofit_columns(sheet, headers: tuple[str, ...]) -> None:
    widest = [len(header) for header in headers]
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            length = _display_length(cell.value)
            index = cell.column - 1
            while index >= len(widest):
                widest.append(0)
            if length > widest[index]:
                widest[index] = length
    for index, width in enumerate(widest, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = (
            width + _COLUMN_WIDTH_PADDING
        )
        if sheet.column_dimensions[get_column_letter(index)].width < _MIN_COLUMN_WIDTH:
            sheet.column_dimensions[get_column_letter(index)].width = _MIN_COLUMN_WIDTH


def _uniform(values: list[str]) -> str:
    unique = sorted({value.strip() for value in values if value.strip()})
    return unique[0] if len(unique) == 1 else ""


def _litter_count(chunk: list[SourceMouse]) -> int:
    seen: set[tuple[str, int]] = set()
    total = 0
    for mouse in chunk:
        if mouse.source_cage_key not in seen:
            seen.add(mouse.source_cage_key)
            total += mouse.source_in_litter
    return total


def _date_bounds(chunk: list[SourceMouse]) -> tuple[date | None, date | None]:
    lows = [mouse.dob_min for mouse in chunk if mouse.dob_min is not None]
    highs = [mouse.dob_max for mouse in chunk if mouse.dob_max is not None]
    if not lows or not highs:
        return None, None
    return min(lows), max(highs)


def _date_value(low: date | None, high: date | None) -> date | str:
    if low is None or high is None:
        return ""
    if low == high:
        return low
    return f"{low.strftime('%m/%d/%y')} - {high.strftime('%m/%d/%y')}"


def _row_for_chunk(chunk: list[SourceMouse]) -> list:
    experiment_url = _uniform([mouse.experiment_url for mouse in chunk])
    strain = chunk[0].strain.strip()
    sex = _uniform([mouse.sex for mouse in chunk]) or chunk[0].sex
    dob_low, dob_high = _date_bounds(chunk)
    date_born = _date_value(dob_low, dob_high)
    if dob_low is not None and dob_high is not None:
        date_weaned = _date_value(dob_low + timedelta(days=28), dob_high + timedelta(days=28))
    else:
        date_weaned = ""
    dam = _uniform([mouse.dam for mouse in chunk])
    dam_genotype = _uniform([mouse.dam_genotype for mouse in chunk]) if dam else ""
    sire = _uniform([mouse.sire for mouse in chunk])
    sire_genotype = _uniform([mouse.sire_genotype for mouse in chunk]) if sire else ""
    breeder = _uniform([mouse.breeder for mouse in chunk])

    mouse_ids = ([mouse.mouse_id for mouse in chunk] + [""] * MAX_MICE_PER_CAGE)[:MAX_MICE_PER_CAGE]
    genotypes = ([mouse.genotype for mouse in chunk] + [""] * MAX_MICE_PER_CAGE)[:MAX_MICE_PER_CAGE]

    return [
        experiment_url,
        strain,
        len(chunk),
        _litter_count(chunk),
        sex,
        date_weaned,
        date_born,
        *mouse_ids,
        *genotypes,
        dam,
        dam_genotype,
        sire,
        sire_genotype,
        breeder,
        "",
    ]


def _group_by_source_cage(mice: list[SourceMouse]) -> list[list[SourceMouse]]:
    groups: dict[tuple[str, int], list[SourceMouse]] = {}
    order: list[tuple[str, int]] = []
    for mouse in mice:
        key = mouse.source_cage_key
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(mouse)
    return [groups[key] for key in order]


def build_consolidated_workbook(
    result: ConsolidationResult,
    output_path: Path,
    *,
    input_reports: list[dict] | None = None,
) -> Path:
    """Write the consolidated cages to ``Sheet1``, preserved-but-unconsolidated
    cage rows to a separate ``Unconsolidated`` sheet, per-mouse review detail
    to a ``Review Needed`` sheet, and a small run report to a ``Report``
    sheet — all in one new Live Label-format workbook.

    Unconsolidated mice are deliberately kept out of ``Sheet1`` (which used
    to blend them in, indistinguishable from a real consolidated cage) so a
    reader can never mistake a single preserved mouse for a successfully
    merged group.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(EXPECTED_HEADERS)
    for cage in result.consolidated_cages:
        sheet.append(_row_for_chunk(cage.mice))
    _autofit_columns(sheet, EXPECTED_HEADERS)

    unconsolidated_sheet = workbook.create_sheet("Unconsolidated")
    unconsolidated_sheet.append(EXPECTED_HEADERS)
    for group in _group_by_source_cage(result.unconsolidated_mice):
        unconsolidated_sheet.append(_row_for_chunk(group))
    _autofit_columns(unconsolidated_sheet, EXPECTED_HEADERS)

    review_sheet = workbook.create_sheet("Review Needed")
    review_sheet.append(_REVIEW_HEADERS)
    for mouse in result.unconsolidated_mice:
        review_sheet.append(
            [
                mouse.source_file,
                mouse.source_row,
                mouse.mouse_id,
                mouse.strain,
                mouse.sex,
                mouse.genotype,
                "; ".join(unconsolidated_reasons(mouse)),
            ]
        )
    _autofit_columns(review_sheet, _REVIEW_HEADERS)

    report_sheet = workbook.create_sheet("Report")
    report_rows = [
        ("Kras genotype grammar version", KRAS_GENOTYPE_GRAMMAR_VERSION),
        ("Input cage count", result.input_cage_count),
        ("Input mouse count", result.input_mouse_count),
        ("Consolidated cage count", len(result.consolidated_cages)),
        (
            "Preserved (unconsolidated) cage count",
            len(_group_by_source_cage(result.unconsolidated_mice)),
        ),
        ("Unconsolidated mouse count", len(result.unconsolidated_mice)),
        ("Warning count", len(result.warnings)),
    ]
    for label, value in report_rows:
        report_sheet.append([label, value])
    if input_reports:
        report_sheet.append([])
        report_sheet.append(["Input File", "SHA-256", "Row Count"])
        for entry in input_reports:
            report_sheet.append(
                [entry.get("filename", ""), entry.get("sha256", ""), entry.get("row_count", "")]
            )
    if result.warnings:
        report_sheet.append([])
        report_sheet.append(["Warnings"])
        for warning in result.warnings:
            report_sheet.append([warning])
    _autofit_columns(report_sheet, ("Label", "Value"))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        raise FileExistsError(f"Refusing to overwrite existing output: {output_path}")
    workbook.save(output_path)
    return output_path
