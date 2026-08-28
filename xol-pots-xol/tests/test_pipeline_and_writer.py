from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from xolpotsxol.models import EXPECTED_HEADERS
from xolpotsxol.pipeline import run_consolidation


def _build_workbook(path: Path, rows: list[list]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(EXPECTED_HEADERS)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


class RunConsolidationTests(unittest.TestCase):
    def test_cross_file_consolidation_writes_expected_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            file_a = root / "a.xlsx"
            file_b = root / "b.xlsx"
            _build_workbook(
                file_a,
                [
                    [
                        "", "Kras/Lkb1", 1, 1, "Male", "02/16/26", "01/19/26",
                        "CM0001", "", "", "", "", "+/+", "", "", "", "",
                        "", "", "", "", "", "",
                    ],
                ],
            )
            _build_workbook(
                file_b,
                [
                    [
                        "", "Kras/Lkb1", 1, 1, "Male", "02/17/26", "01/20/26",
                        "CM0002", "", "", "", "", "+/+", "", "", "", "",
                        "", "", "", "", "", "",
                    ],
                ],
            )
            output_path = root / "out.xlsx"

            result, read_warnings = run_consolidation(
                [(file_a, "a.xlsx"), (file_b, "b.xlsx")], output_path
            )

            self.assertEqual(read_warnings, [])
            self.assertEqual(result.input_cage_count, 2)
            self.assertEqual(result.input_mouse_count, 2)
            self.assertEqual(len(result.consolidated_cages), 1)
            self.assertTrue(output_path.is_file())

            workbook = load_workbook(output_path)
            sheet = workbook.active
            self.assertEqual(
                tuple(sheet.cell(1, c).value for c in range(1, 24)), EXPECTED_HEADERS
            )
            data_row = [sheet.cell(2, c).value for c in range(1, 24)]
            self.assertEqual(data_row[2], 2)  # # IN CAGE
            self.assertEqual(data_row[3], 2)  # # IN LITTER: 1 + 1 from two distinct source rows
            self.assertEqual(set(data_row[7:9]), {"CM0001", "CM0002"})

    def test_refuses_to_overwrite_existing_output(self) -> None:
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            file_a = root / "a.xlsx"
            _build_workbook(
                file_a,
                [
                    [
                        "", "Kras/Lkb1", 1, 1, "Male", "02/16/26", "01/19/26",
                        "CM0001", "", "", "", "", "+/+", "", "", "", "",
                        "", "", "", "", "", "",
                    ],
                ],
            )
            output_path = root / "out.xlsx"
            output_path.write_text("existing", encoding="utf-8")

            with self.assertRaises(FileExistsError):
                run_consolidation([(file_a, "a.xlsx")], output_path)


if __name__ == "__main__":
    unittest.main()
