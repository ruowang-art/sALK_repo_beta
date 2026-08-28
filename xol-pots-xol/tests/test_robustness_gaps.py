"""Regression tests for gaps surfaced by an external robustness review of
Xol-Pots-Xol (see mouseley-kras-and-xol-pots-xol-overview.md in the sibling
Möuseley Kräs repo): an unrecognized Kras genotype flowing all the way
through consolidate(), a reordered header row, duplicate mouse IDs, input
immutability, and the new Unconsolidated/Review Needed/Report worksheets.
"""

from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from xolpotsxol.consolidator import consolidate, unconsolidated_reasons
from xolpotsxol.exceptions import CageCardFormatError
from xolpotsxol.models import EXPECTED_HEADERS, SourceMouse
from xolpotsxol.cage_card_reader import read_cage_card_workbook
from xolpotsxol.pipeline import run_consolidation


def _mouse(
    mouse_id: str,
    *,
    sex: str = "Male",
    strain: str = "Kras/Lkb1",
    genotype: str = "+/+",
    dob: date | None = date(2026, 1, 1),
    source_row: int = 1,
    source_file: str = "a.xlsx",
) -> SourceMouse:
    return SourceMouse(
        mouse_id=mouse_id,
        genotype=genotype,
        sex=sex,
        strain=strain,
        dob_min=dob,
        dob_max=dob,
        dam="", dam_genotype="", sire="", sire_genotype="", breeder="",
        experiment_url="", source_file=source_file, source_row=source_row,
        source_in_litter=1,
    )


def _build_workbook(path: Path, rows: list[list], headers: tuple = EXPECTED_HEADERS) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


class UnrecognizedGenotypeEndToEndTests(unittest.TestCase):
    def test_unrecognized_kras_string_stays_unconsolidated(self) -> None:
        mice = [_mouse("CM0001", genotype="weird"), _mouse("CM0002", genotype="+/+")]
        result = consolidate(mice)
        self.assertEqual([m.mouse_id for m in result.unconsolidated_mice], ["CM0001"])
        self.assertEqual(len(result.consolidated_cages), 1)

    def test_unconsolidated_reasons_names_the_bad_genotype(self) -> None:
        mouse = _mouse("CM0001", genotype="weird")
        reasons = unconsolidated_reasons(mouse)
        self.assertTrue(any("weird" in reason for reason in reasons))

    def test_unconsolidated_reasons_can_report_multiple_causes(self) -> None:
        mouse = _mouse("CM0001", genotype="weird", sex="", strain="")
        reasons = unconsolidated_reasons(mouse)
        self.assertEqual(len(reasons), 3)


class ReorderedHeaderTests(unittest.TestCase):
    def test_reordered_header_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            path = root / "reordered.xlsx"
            reordered = (EXPECTED_HEADERS[1], EXPECTED_HEADERS[0]) + EXPECTED_HEADERS[2:]
            _build_workbook(path, [], headers=reordered)

            with self.assertRaises(CageCardFormatError):
                read_cage_card_workbook(path, "reordered.xlsx")


class DuplicateMouseIdTests(unittest.TestCase):
    def test_duplicate_mouse_id_within_one_workbook_is_read_as_two_rows(self) -> None:
        # Xol-Pots-Xol has no inventory to check identity against (unlike
        # Möuseley Kräs); it currently just reads whatever rows are present.
        # This test documents that behavior rather than changing it.
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            path = root / "dupe.xlsx"
            row = [
                "", "Kras/Lkb1", 1, 1, "Male", "01/01/26", "01/01/26",
                "CM0001", "", "", "", "", "+/+", "", "", "", "",
                "", "", "", "", "", "",
            ]
            _build_workbook(path, [row, row])

            mice, warnings = read_cage_card_workbook(path, "dupe.xlsx")

            self.assertEqual([m.mouse_id for m in mice], ["CM0001", "CM0001"])
            self.assertEqual(warnings, [])


class InputImmutabilityTests(unittest.TestCase):
    def test_input_workbooks_are_unchanged_after_a_successful_run(self) -> None:
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            path = root / "a.xlsx"
            _build_workbook(
                path,
                [[
                    "", "Kras/Lkb1", 1, 1, "Male", "01/01/26", "01/01/26",
                    "CM0001", "", "", "", "", "+/+", "", "", "", "",
                    "", "", "", "", "", "",
                ]],
            )
            before_bytes = path.read_bytes()
            before_mtime = path.stat().st_mtime_ns

            run_consolidation([(path, "a.xlsx")], root / "out.xlsx")

            self.assertEqual(path.read_bytes(), before_bytes)
            self.assertEqual(path.stat().st_mtime_ns, before_mtime)

    def test_input_workbooks_are_unchanged_after_a_failed_run(self) -> None:
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            path = root / "bad.xlsx"
            _build_workbook(path, [], headers=("Not", "The", "Right", "Headers"))
            before_bytes = path.read_bytes()

            with self.assertRaises(CageCardFormatError):
                run_consolidation([(path, "bad.xlsx")], root / "out.xlsx")

            self.assertEqual(path.read_bytes(), before_bytes)


class OutputWorksheetTests(unittest.TestCase):
    def test_output_has_separate_unconsolidated_review_and_report_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            path = root / "a.xlsx"
            _build_workbook(
                path,
                [
                    [
                        "", "Kras/Lkb1", 1, 1, "Male", "01/01/26", "01/01/26",
                        "CM0001", "", "", "", "", "weird", "", "", "", "",
                        "", "", "", "", "", "",
                    ]
                ],
            )
            output_path = root / "out.xlsx"

            result, _ = run_consolidation([(path, "a.xlsx")], output_path)

            self.assertEqual(len(result.unconsolidated_mice), 1)
            workbook = load_workbook(output_path)
            self.assertEqual(
                set(workbook.sheetnames),
                {"Sheet1", "Unconsolidated", "Review Needed", "Report"},
            )

            # The unrecognized-genotype mouse must NOT appear in Sheet1
            # (the consolidated-only sheet), only in Unconsolidated.
            sheet1_rows = list(workbook["Sheet1"].iter_rows(min_row=2, values_only=True))
            self.assertEqual(sheet1_rows, [])
            unconsolidated_rows = list(
                workbook["Unconsolidated"].iter_rows(min_row=2, values_only=True)
            )
            self.assertEqual(len(unconsolidated_rows), 1)
            self.assertIn("CM0001", unconsolidated_rows[0])

            review_rows = list(workbook["Review Needed"].iter_rows(min_row=2, values_only=True))
            self.assertEqual(len(review_rows), 1)
            self.assertEqual(review_rows[0][2], "CM0001")  # Mouse ID column
            self.assertIn("weird", review_rows[0][5])  # Raw Genotype column
            self.assertTrue(review_rows[0][6])  # Reason(s) column is non-empty

            report_values = [
                cell for row in workbook["Report"].iter_rows(values_only=True) for cell in row
            ]
            self.assertIn("a.xlsx", report_values)


if __name__ == "__main__":
    unittest.main()
