from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from xolpotsxol.models import ConsolidatedCage, ConsolidationResult, EXPECTED_HEADERS, SourceMouse
from xolpotsxol.writer import build_consolidated_workbook


def _mouse(mouse_id: str, dob_min: date, dob_max: date, genotype: str) -> SourceMouse:
    return SourceMouse(
        mouse_id=mouse_id,
        genotype=genotype,
        sex="Male",
        strain="Kras/Lkb1/Tom/Cas9",
        dob_min=dob_min,
        dob_max=dob_max,
        dam="",
        dam_genotype="",
        sire="",
        sire_genotype="",
        breeder="",
        experiment_url="",
        source_file="a.xlsx",
        source_row=1,
        source_in_litter=1,
    )


class ColumnWidthTests(unittest.TestCase):
    def test_every_column_width_fits_its_actual_longest_value(self) -> None:
        # A wide DOB range and a long composite genotype string are exactly
        # the kind of content that showed up as "####" in Excel when column
        # widths were too narrow (or guessed wrong) for a workbook built
        # from scratch.
        cage = ConsolidatedCage(
            mice=[
                _mouse("CM0001", date(2026, 1, 1), date(2026, 1, 1), "+/+; L/L; T/T; HC9/HC9"),
                _mouse("CM0002", date(2026, 1, 8), date(2026, 1, 8), "+/+; L/L; T/T; HC9/HC9"),
            ],
            warnings=[],
        )
        result = ConsolidationResult(
            consolidated_cages=[cage],
            unconsolidated_mice=[],
            warnings=[],
            input_cage_count=2,
            input_mouse_count=2,
        )

        with tempfile.TemporaryDirectory() as directory_name:
            output_path = Path(directory_name) / "out.xlsx"
            build_consolidated_workbook(result, output_path)
            workbook = load_workbook(output_path)
            sheet = workbook.active

            for column_index, header in enumerate(EXPECTED_HEADERS, start=1):
                longest = len(header)
                for row_index in range(2, sheet.max_row + 1):
                    value = sheet.cell(row_index, column_index).value
                    if value is None or value == "":
                        continue
                    text = value.strftime("%Y-%m-%d") if hasattr(value, "strftime") else str(value)
                    longest = max(longest, len(text))
                width = sheet.column_dimensions[get_column_letter(column_index)].width
                self.assertGreaterEqual(
                    width,
                    longest,
                    f"column {header!r} width {width} is narrower than its longest value ({longest} chars)",
                )


if __name__ == "__main__":
    unittest.main()
