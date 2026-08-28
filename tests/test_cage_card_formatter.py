from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side

from automouse.cage_card_formatter import generate_cage_cards
from automouse.config import CageCardConfig
from automouse.models import CageCardRecord

TEMPLATE_HEADERS = (
    "Experiment URL", "Strain", "# IN CAGE", "# IN LITTER", "SEX",
    "DATE WEANED", "DATE BORN", "MOUSE 1", "MOUSE 2", "MOUSE 3",
    "MOUSE 4", "MOUSE 5", "MOUSE 1 GENOTYPE", "MOUSE 2 GENOTYPE",
    "MOUSE 3 GENOTYPE", "MOUSE 4 GENOTYPE", "MOUSE 5 GENOTYPE",
    "DAM", "DAM GENOTYPE", "SIRE", "SIRE GENOTYPE", "BREEDER? (B)",
    "Set up date",
)


class GenerateCageCardsBorderTests(unittest.TestCase):
    def _build_template_with_stray_borders(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(TEMPLATE_HEADERS)
        for _ in range(70):
            sheet.append([""] * len(TEMPLATE_HEADERS))
        # Stray leftover formatting far below any real card row, matching
        # what was found in the real Live Label template.
        thin = Border(left=Side(style="thin"))
        sheet.cell(20, 13).border = thin
        sheet.cell(53, 18).border = thin
        workbook.save(path)

    def test_blank_rows_below_the_populated_cards_have_no_border(self) -> None:
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            template_path = root / "template.xlsx"
            self._build_template_with_stray_borders(template_path)

            output_path = root / "cage_cards.xlsx"
            record = CageCardRecord(
                cage_id="Female|Kras|+/+|2026-01-01",
                experiment_url="",
                strain="Kras",
                animal_count=1,
                litter_count=1,
                sex="Female",
                date_weaned=None,
                date_born=None,
                mouse_ids=["CM0001"],
                genotypes=["+/+"],
                dam="",
                dam_genotype="",
                sire="",
                sire_genotype="",
            )
            config = CageCardConfig(template=template_path, expected_headers=TEMPLATE_HEADERS)

            generate_cage_cards([record], template_path, output_path, config)

            workbook = load_workbook(output_path)
            sheet = workbook["Sheet1"]
            for row in (20, 53):
                for column in range(1, 24):
                    border = sheet.cell(row, column).border
                    for side in (border.left, border.right, border.top, border.bottom):
                        self.assertIsNone(
                            side.style if side is not None else None,
                            f"row {row} col {column}",
                        )


if __name__ == "__main__":
    unittest.main()
