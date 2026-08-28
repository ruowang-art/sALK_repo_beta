from __future__ import annotations

import json
import csv
import hashlib
import sys
import tempfile
import types
import unittest
from pathlib import Path
from unittest import mock
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from automouse.app import run_batch, run_phase_1_2
from automouse.config import (
    AppConfig,
    CageCardConfig,
    InventoryConfig,
    RConfig,
    SheetsOverlayConfig,
    TransnetyxConfig,
    TranslationConfig,
)
from automouse.models import RRunResult, RunStage


FIXTURES = Path(__file__).parent / "fixtures"


class AppIntegrationTests(unittest.TestCase):
    def test_phase_1_2_creates_expected_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            executable = root / "Rscript"
            translation = root / "translation.R"
            wrapper = root / "wrapper.R"
            for path in (executable, translation, wrapper):
                path.write_text("fixture", encoding="utf-8")
            config = AppConfig(
                project_root=root,
                runtime_root=root / "runtime",
                application_version="test",
                r=RConfig(executable, translation, wrapper),
                transnetyx=TransnetyxConfig(),
                translation=TranslationConfig(),
            )

            def fake_translation(input_path: Path, output_path: Path, *_: object, **__: object) -> RRunResult:
                output_path.write_bytes((FIXTURES / "translated_valid.csv").read_bytes())
                return RRunResult([], 0, "", "", output_path, 0.01)

            with patch("automouse.app.run_r_translation", side_effect=fake_translation):
                context = run_phase_1_2(FIXTURES / "raw_valid.csv", config, dry_run=True)

            self.assertEqual(context.stage, RunStage.COMPLETED)
            self.assertTrue(Path(context.artifacts["archived_input_file"]).is_file())
            self.assertTrue(Path(context.artifacts["translated_output_file"]).is_file())
            self.assertTrue(Path(context.artifacts["translation_validation_file"]).is_file())
            summary_path = Path(context.artifacts["run_summary_file"])
            self.assertTrue(summary_path.is_file())
            summary = json.loads(summary_path.read_text(encoding="utf-8"))
            self.assertEqual(summary["raw_record_count"], 2)
            self.assertEqual(summary["translated_record_count"], 2)
            self.assertTrue(summary["dry_run"])

    def test_complete_pipeline_updates_copy_and_generates_cage_card(self) -> None:
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            executable = root / "Rscript"
            translation = root / "translation.R"
            wrapper = root / "wrapper.R"
            for path in (executable, translation, wrapper):
                path.write_text("fixture", encoding="utf-8")

            inventory_path = root / "master.csv"
            headers = [f"Column {index}" for index in range(1, 22)]
            for index, value in {
                1: "Mouse",
                4: "Strain",
                5: "Cage",
                6: "Revised Strain",
                13: "ID",
                14: "Mother",
                15: "Father",
                16: "DOB",
                17: "Sex",
                18: "Wean_By",
                20: "Sample",
                21: "Genotype",
            }.items():
                headers[index - 1] = value
            rows = []
            for mouse, sex, genotype in (
                ("CM0001", "Male", ""),
                ("CM0002", "Male", "LSL-G12D/+"),
            ):
                row = [""] * 21
                row[0] = mouse
                row[3] = "Kras"
                row[4] = ""
                row[12] = mouse
                row[13] = "CM9001"
                row[14] = "CM9002"
                row[15] = "07/01/2026"
                row[16] = sex
                row[17] = "07/22/2026"
                row[19] = mouse
                row[20] = genotype
                rows.append(row)
            with inventory_path.open("w", encoding="utf-8", newline="") as stream:
                writer = csv.writer(stream)
                writer.writerow(headers)
                writer.writerows(rows)

            template_path = root / "template.xlsx"
            template_headers = (
                "Experiment URL", "Strain", "# IN CAGE", "# IN LITTER", "SEX",
                "DATE WEANED", "DATE BORN", "MOUSE 1", "MOUSE 2", "MOUSE 3",
                "MOUSE 4", "MOUSE 5", "MOUSE 1 GENOTYPE", "MOUSE 2 GENOTYPE",
                "MOUSE 3 GENOTYPE", "MOUSE 4 GENOTYPE", "MOUSE 5 GENOTYPE",
                "DAM", "DAM GENOTYPE", "SIRE", "SIRE GENOTYPE", "BREEDER? (B)",
                "Set up date",
            )
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(template_headers)
            sheet.append([""] * 23)
            workbook.save(template_path)

            inventory_config = InventoryConfig(
                file=inventory_path,
                columns={
                    "mouse_id": 1, "strain": 4, "cage_id": 5,
                    "revised_strain": 6, "id": 13, "mother": 14,
                    "father": 15, "dob": 16, "sex": 17, "wean_date": 18,
                    "sample": 20, "genotype": 21,
                },
                expected_headers={
                    "mouse_id": "Mouse", "strain": "Strain", "cage_id": "Cage",
                    "revised_strain": "Revised Strain", "id": "ID",
                    "mother": "Mother", "father": "Father", "dob": "DOB",
                    "sex": "Sex", "wean_date": "Wean_By", "sample": "Sample",
                    "genotype": "Genotype",
                },
            )
            config = AppConfig(
                project_root=root,
                runtime_root=root / "runtime",
                application_version="test",
                r=RConfig(executable, translation, wrapper),
                transnetyx=TransnetyxConfig(),
                translation=TranslationConfig(),
                inventory=inventory_config,
                cage_card=CageCardConfig(
                    template=template_path,
                    expected_headers=template_headers,
                ),
            )

            raw_one = root / "raw one.csv"
            raw_two = root / "raw two.csv"
            raw_header = "WellPlate,Strain,Well,Sample,TranslatedResult,G12D mut,Y\n"
            raw_one.write_text(
                raw_header + "TEST001,Kras,A1,CM0001,,--,+\n",
                encoding="utf-8",
            )
            raw_two.write_text(
                raw_header + "TEST002,Kras,A2,CM0002,,+-, -\n",
                encoding="utf-8",
            )
            translated_rows = iter(
                [
                    "TEST001,Kras,A1,CM0001,,--,+,+/+,Male,'+/+\n",
                    "TEST002,Kras,A2,CM0002,,+-, -,LSL-G12D/+,Male,LSL-G12D/+\n",
                ]
            )

            def fake_translation(input_path: Path, output_path: Path, *_: object, **__: object) -> RRunResult:
                output_path.write_text(
                    "WellPlate,Strain,Well,Sample,TranslatedResult,G12D.mut,Y,Kras,Sex,Genotype\n"
                    + next(translated_rows),
                    encoding="utf-8",
                )
                return RRunResult([], 0, "", "", output_path, 0.01)

            original_bytes = inventory_path.read_bytes()
            with patch("automouse.app.run_r_translation", side_effect=fake_translation):
                context = run_batch(
                    [raw_one, raw_two],
                    config,
                    complete_pipeline=True,
                )

            self.assertEqual(context.stage, RunStage.COMPLETED)
            self.assertEqual(context.counts["inventory_records_updated"], 1)
            self.assertEqual(context.counts["inventory_records_confirmed"], 1)
            self.assertEqual(context.counts["cage_cards_generated"], 2)
            self.assertEqual(context.counts["input_file_count"], 2)
            self.assertEqual(context.counts["card_eligible_count"], 2)
            self.assertEqual(inventory_path.read_bytes(), original_bytes)
            self.assertEqual(
                Path(context.artifacts["inventory_backup_file"]).read_bytes(),
                original_bytes,
            )

            updated = load_workbook(context.artifacts["updated_inventory_file"], read_only=True)
            self.assertEqual(updated["Mouse Inventory"].cell(2, 21).value, "+/+")
            updated.close()
            cards = load_workbook(context.artifacts["cage_card_file"], read_only=True)
            self.assertEqual(cards["Sheet1"].cell(2, 3).value, 1)
            self.assertEqual(cards["Sheet1"].cell(2, 8).value, "CM0001")
            self.assertEqual(cards["Sheet1"].cell(3, 8).value, "CM0002")
            cards.close()

            with Path(context.artifacts["audit_file"]).open(newline="") as stream:
                audit_rows = list(csv.DictReader(stream))
            self.assertEqual(len({row["source_file"] for row in audit_rows}), 2)
            index = json.loads(
                (config.runtime_root / "archive/raw_transnetyx_files/checksum_index.json")
                .read_text(encoding="utf-8")
            )
            run_entries = [
                entry for entry in index["entries"] if entry["run_id"] == context.run_id
            ]
            self.assertEqual(len(run_entries), 2)
            self.assertEqual({entry["status"] for entry in run_entries}, {"completed"})

    def test_append_only_blank_inventory_builds_new_rows(self) -> None:
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            executable = root / "Rscript"
            translation = root / "translation.R"
            wrapper = root / "wrapper.R"
            for path in (executable, translation, wrapper):
                path.write_text("fixture", encoding="utf-8")

            inventory_path = root / "test_mouse_inventory.csv"
            headers = [f"Column {index}" for index in range(1, 46)]
            for index, value in {
                1: "Mouse",
                4: "Strain",
                5: "Cage",
                6: "Revised Strain",
                13: "ID",
                14: "Mother",
                15: "Father",
                16: "DOB",
                17: "Sex",
                18: "Wean_By",
                20: "Sample",
                21: "Genotype",
            }.items():
                headers[index - 1] = value
            with inventory_path.open("w", encoding="utf-8", newline="") as stream:
                writer = csv.writer(stream)
                writer.writerow(headers)

            template_path = root / "template.xlsx"
            template_headers = (
                "Experiment URL", "Strain", "# IN CAGE", "# IN LITTER", "SEX",
                "DATE WEANED", "DATE BORN", "MOUSE 1", "MOUSE 2", "MOUSE 3",
                "MOUSE 4", "MOUSE 5", "MOUSE 1 GENOTYPE", "MOUSE 2 GENOTYPE",
                "MOUSE 3 GENOTYPE", "MOUSE 4 GENOTYPE", "MOUSE 5 GENOTYPE",
                "DAM", "DAM GENOTYPE", "SIRE", "SIRE GENOTYPE", "BREEDER? (B)",
                "Set up date",
            )
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(template_headers)
            workbook.save(template_path)

            inventory_config = InventoryConfig(
                file=inventory_path,
                append_only=True,
                columns={
                    "mouse_id": 1, "strain": 4, "cage_id": 5,
                    "revised_strain": 6, "id": 13, "mother": 14,
                    "father": 15, "dob": 16, "sex": 17, "wean_date": 18,
                    "sample": 20, "genotype": 21,
                },
                expected_headers={
                    "mouse_id": "Mouse", "strain": "Strain", "cage_id": "Cage",
                    "revised_strain": "Revised Strain", "id": "ID",
                    "mother": "Mother", "father": "Father", "dob": "DOB",
                    "sex": "Sex", "wean_date": "Wean_By", "sample": "Sample",
                    "genotype": "Genotype",
                },
            )
            config = AppConfig(
                project_root=root,
                runtime_root=root / "runtime",
                application_version="test",
                r=RConfig(executable, translation, wrapper),
                transnetyx=TransnetyxConfig(),
                translation=TranslationConfig(),
                inventory=inventory_config,
                cage_card=CageCardConfig(
                    template=template_path,
                    expected_headers=template_headers,
                ),
            )

            raw_one = root / "raw one.csv"
            raw_two = root / "raw two.csv"
            raw_one.write_text(
                "WellPlate,Strain,Well,Sample,Sex,Genotype\n"
                "T1,Kras/p53/Tom/Cas9/Cpt2,A1,CM9001,Male,+/+\n",
                encoding="utf-8",
            )
            raw_two.write_text(
                "WellPlate,Strain,Well,Sample,Sex,Genotype\n"
                "T2,Kras/p53/Tom/Cas9/Cpt2,A2,CM9002,Female,K/+\n",
                encoding="utf-8",
            )

            translated_rows = {
                "raw one": (
                    "WellPlate,Strain,Well,Sample,Sex,Genotype\n"
                    "T1,Kras/p53/Tom/Cas9/Cpt2,A1,CM9001,Male,+/+\n"
                ),
                "raw two": (
                    "WellPlate,Strain,Well,Sample,Sex,Genotype\n"
                    "T2,Kras/p53/Tom/Cas9/Cpt2,A2,CM9002,Female,K/+\n"
                ),
            }

            def fake_translation(input_path: Path, output_path: Path, *_: object, **__: object) -> RRunResult:
                key = "raw one" if ("raw one" in input_path.name or "raw_one" in input_path.name) else "raw two"
                output_path.write_text(translated_rows[key], encoding="utf-8")
                return RRunResult([], 0, "", "", output_path, 0.01)

            with patch("automouse.app.run_r_translation", side_effect=fake_translation):
                context = run_batch(
                    [raw_one, raw_two],
                    config,
                    complete_pipeline=True,
                )

            self.assertEqual(context.stage, RunStage.COMPLETED)
            self.assertEqual(context.counts["inventory_record_count"], 0)
            self.assertEqual(context.counts["inventory_records_updated"], 2)
            self.assertEqual(context.counts["inventory_records_confirmed"], 0)
            self.assertEqual(context.counts["cage_cards_generated"], 2)
            self.assertTrue(Path(context.artifacts["cage_card_file"]).is_file())

            with inventory_path.open(newline="", encoding="utf-8-sig") as stream:
                inventory_rows = list(csv.DictReader(stream))
            self.assertEqual(len(inventory_rows), 2)
            self.assertEqual(inventory_rows[0]["Mouse"], "CM9001")
            self.assertEqual(inventory_rows[0]["Strain"], "Kras/p53/Tom/Cas9/Cpt2")
            self.assertEqual(inventory_rows[0]["Revised Strain"], "Kras/p53/Tom/Cas9/Cpt2")
            self.assertEqual(inventory_rows[0]["Sex"], "Male")
            self.assertEqual(inventory_rows[0]["Sample"], "CM9001")
            self.assertEqual(inventory_rows[0]["Genotype"], "+/+")
            self.assertEqual(inventory_rows[1]["Mouse"], "CM9002")

            updated = load_workbook(context.artifacts["updated_inventory_file"], read_only=True)
            sheet = updated["Mouse Inventory"]
            self.assertEqual(sheet.cell(2, 1).value, "CM9001")
            self.assertEqual(sheet.cell(2, 4).value, "Kras/p53/Tom/Cas9/Cpt2")
            self.assertEqual(sheet.cell(2, 6).value, "Kras/p53/Tom/Cas9/Cpt2")
            self.assertEqual(sheet.cell(2, 17).value, "Male")
            self.assertEqual(sheet.cell(2, 20).value, "CM9001")
            self.assertEqual(sheet.cell(2, 21).value, "+/+")
            updated.close()

            cards = load_workbook(context.artifacts["cage_card_file"], read_only=True)
            card_sheet = cards["Sheet1"]
            self.assertEqual(card_sheet.cell(2, 8).value, "CM9001")
            self.assertEqual(card_sheet.cell(3, 8).value, "CM9002")
            self.assertIsNone(card_sheet.cell(2, 6).value)
            self.assertIsNone(card_sheet.cell(2, 7).value)
            cards.close()

            with patch("automouse.app.run_r_translation", side_effect=fake_translation):
                rerun_context = run_batch(
                    [raw_one, raw_two],
                    config,
                    allow_duplicate_input=True,
                    complete_pipeline=True,
                )

            self.assertEqual(rerun_context.stage, RunStage.COMPLETED)
            self.assertEqual(rerun_context.counts["inventory_records_updated"], 0)
            self.assertEqual(rerun_context.counts["cage_cards_generated"], 2)

    def test_dry_run_over_new_append_only_mice_warns_instead_of_previewing_cards(self) -> None:
        # A dry run never appends the new inventory row an append-only config
        # would otherwise gain, so there is no row yet to build a cage card
        # from. Without an explicit warning this looks exactly like a silent
        # failure: 0 cage cards with no explanation.
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            executable = root / "Rscript"
            translation = root / "translation.R"
            wrapper = root / "wrapper.R"
            for path in (executable, translation, wrapper):
                path.write_text("fixture", encoding="utf-8")

            inventory_path = root / "test_mouse_inventory.csv"
            headers = [f"Column {index}" for index in range(1, 22)]
            for index, value in {
                1: "Mouse", 4: "Strain", 5: "Cage", 6: "Revised Strain",
                13: "ID", 14: "Mother", 15: "Father", 16: "DOB", 17: "Sex",
                18: "Wean_By", 20: "Sample", 21: "Genotype",
            }.items():
                headers[index - 1] = value
            with inventory_path.open("w", encoding="utf-8", newline="") as stream:
                writer = csv.writer(stream)
                writer.writerow(headers)

            template_path = root / "template.xlsx"
            template_headers = (
                "Experiment URL", "Strain", "# IN CAGE", "# IN LITTER", "SEX",
                "DATE WEANED", "DATE BORN", "MOUSE 1", "MOUSE 2", "MOUSE 3",
                "MOUSE 4", "MOUSE 5", "MOUSE 1 GENOTYPE", "MOUSE 2 GENOTYPE",
                "MOUSE 3 GENOTYPE", "MOUSE 4 GENOTYPE", "MOUSE 5 GENOTYPE",
                "DAM", "DAM GENOTYPE", "SIRE", "SIRE GENOTYPE", "BREEDER? (B)",
                "Set up date",
            )
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(template_headers)
            workbook.save(template_path)

            config = AppConfig(
                project_root=root,
                runtime_root=root / "runtime",
                application_version="test",
                r=RConfig(executable, translation, wrapper),
                transnetyx=TransnetyxConfig(),
                translation=TranslationConfig(),
                inventory=InventoryConfig(
                    file=inventory_path,
                    append_only=True,
                    columns={
                        "mouse_id": 1, "strain": 4, "cage_id": 5,
                        "revised_strain": 6, "id": 13, "mother": 14,
                        "father": 15, "dob": 16, "sex": 17, "wean_date": 18,
                        "sample": 20, "genotype": 21,
                    },
                    expected_headers={
                        "mouse_id": "Mouse", "strain": "Strain", "cage_id": "Cage",
                        "revised_strain": "Revised Strain", "id": "ID",
                        "mother": "Mother", "father": "Father", "dob": "DOB",
                        "sex": "Sex", "wean_date": "Wean_By", "sample": "Sample",
                        "genotype": "Genotype",
                    },
                ),
                cage_card=CageCardConfig(
                    template=template_path,
                    expected_headers=template_headers,
                ),
            )

            raw_one = root / "raw one.csv"
            raw_one.write_text(
                "WellPlate,Strain,Well,Sample,Sex,Genotype\n"
                "T1,Kras/p53/Tom/Cas9/Cpt2,A1,CM9001,Male,+/+\n",
                encoding="utf-8",
            )

            def fake_translation(input_path: Path, output_path: Path, *_: object, **__: object) -> RRunResult:
                output_path.write_text(
                    "WellPlate,Strain,Well,Sample,Sex,Genotype\n"
                    "T1,Kras/p53/Tom/Cas9/Cpt2,A1,CM9001,Male,+/+\n",
                    encoding="utf-8",
                )
                return RRunResult([], 0, "", "", output_path, 0.01)

            with patch("automouse.app.run_r_translation", side_effect=fake_translation):
                context = run_batch(
                    [raw_one],
                    config,
                    dry_run=True,
                    complete_pipeline=True,
                )

            self.assertEqual(context.stage, RunStage.COMPLETED)
            self.assertEqual(context.counts["proposed_update_count"], 1)
            self.assertEqual(context.counts["inventory_records_updated"], 0)
            self.assertEqual(context.counts["card_eligible_count"], 0)
            self.assertEqual(context.counts["cage_cards_generated"], 0)
            self.assertTrue(
                any("preview only" in warning or "dry run" in warning for warning in context.warnings),
                context.warnings,
            )

    def _build_two_mouse_fixture(self, root: Path):
        executable = root / "Rscript"
        translation = root / "translation.R"
        wrapper = root / "wrapper.R"
        for path in (executable, translation, wrapper):
            path.write_text("fixture", encoding="utf-8")

        inventory_path = root / "master.csv"
        headers = [f"Column {index}" for index in range(1, 22)]
        for index, value in {
            1: "Mouse", 4: "Strain", 5: "Cage", 6: "Revised Strain",
            13: "ID", 14: "Mother", 15: "Father", 16: "DOB", 17: "Sex",
            18: "Wean_By", 20: "Sample", 21: "Genotype",
        }.items():
            headers[index - 1] = value
        rows = []
        for mouse, sex, genotype in (
            ("CM0001", "Male", ""),
            ("CM0002", "Male", "LSL-G12D/+"),
        ):
            row = [""] * 21
            row[0] = mouse
            row[3] = "Kras"
            row[12] = mouse
            row[13] = "CM9001"
            row[14] = "CM9002"
            row[15] = "07/01/2026"
            row[16] = sex
            row[17] = "07/22/2026"
            row[19] = mouse
            row[20] = genotype
            rows.append(row)
        with inventory_path.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.writer(stream)
            writer.writerow(headers)
            writer.writerows(rows)

        template_path = root / "template.xlsx"
        template_headers = (
            "Experiment URL", "Strain", "# IN CAGE", "# IN LITTER", "SEX",
            "DATE WEANED", "DATE BORN", "MOUSE 1", "MOUSE 2", "MOUSE 3",
            "MOUSE 4", "MOUSE 5", "MOUSE 1 GENOTYPE", "MOUSE 2 GENOTYPE",
            "MOUSE 3 GENOTYPE", "MOUSE 4 GENOTYPE", "MOUSE 5 GENOTYPE",
            "DAM", "DAM GENOTYPE", "SIRE", "SIRE GENOTYPE", "BREEDER? (B)",
            "Set up date",
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(template_headers)
        sheet.append([""] * 23)
        workbook.save(template_path)

        config = AppConfig(
            project_root=root,
            runtime_root=root / "runtime",
            application_version="test",
            r=RConfig(executable, translation, wrapper),
            transnetyx=TransnetyxConfig(),
            translation=TranslationConfig(),
            inventory=InventoryConfig(
                file=inventory_path,
                columns={
                    "mouse_id": 1, "strain": 4, "cage_id": 5,
                    "revised_strain": 6, "id": 13, "mother": 14,
                    "father": 15, "dob": 16, "sex": 17, "wean_date": 18,
                    "sample": 20, "genotype": 21,
                },
                expected_headers={
                    "mouse_id": "Mouse", "strain": "Strain", "cage_id": "Cage",
                    "revised_strain": "Revised Strain", "id": "ID",
                    "mother": "Mother", "father": "Father", "dob": "DOB",
                    "sex": "Sex", "wean_date": "Wean_By", "sample": "Sample",
                    "genotype": "Genotype",
                },
            ),
            cage_card=CageCardConfig(
                template=template_path,
                expected_headers=template_headers,
            ),
        )

        raw_one = root / "raw one.csv"
        raw_two = root / "raw two.csv"
        raw_header = "WellPlate,Strain,Well,Sample,TranslatedResult,G12D mut,Y\n"
        raw_one.write_text(raw_header + "TEST001,Kras,A1,CM0001,,--,+\n", encoding="utf-8")
        raw_two.write_text(raw_header + "TEST002,Kras,A2,CM0002,,+-, -\n", encoding="utf-8")
        translated_rows = iter(
            [
                "TEST001,Kras,A1,CM0001,,--,+,+/+,Male,'+/+\n",
                "TEST002,Kras,A2,CM0002,,+-, -,LSL-G12D/+,Male,LSL-G12D/+\n",
            ]
        )

        def fake_translation(input_path: Path, output_path: Path, *_: object, **__: object) -> RRunResult:
            output_path.write_text(
                "WellPlate,Strain,Well,Sample,TranslatedResult,G12D.mut,Y,Kras,Sex,Genotype\n"
                + next(translated_rows),
                encoding="utf-8",
            )
            return RRunResult([], 0, "", "", output_path, 0.01)

        return config, raw_one, raw_two, fake_translation, inventory_path

    def _install_fake_google_modules(self, sheet_rows: list[list[str]]) -> None:
        google_module = types.ModuleType("google")
        oauth2_module = types.ModuleType("google.oauth2")
        service_account_module = types.ModuleType("google.oauth2.service_account")

        class _FakeCredentials:
            @classmethod
            def from_service_account_file(cls, path, scopes):
                return cls()

        service_account_module.Credentials = _FakeCredentials
        googleapiclient_module = types.ModuleType("googleapiclient")
        discovery_module = types.ModuleType("googleapiclient.discovery")

        def _build(service_name, version, credentials=None, cache_discovery=True):
            class _Request:
                def execute(self):
                    return {"values": sheet_rows}

            class _Values:
                def get(self, spreadsheetId, range):
                    return _Request()

            class _Spreadsheets:
                def values(self):
                    return _Values()

            class _Service:
                def spreadsheets(self):
                    return _Spreadsheets()

            return _Service()

        discovery_module.build = _build

        self.enterContext(
            mock.patch.dict(
                sys.modules,
                {
                    "google": google_module,
                    "google.oauth2": oauth2_module,
                    "google.oauth2.service_account": service_account_module,
                    "googleapiclient": googleapiclient_module,
                    "googleapiclient.discovery": discovery_module,
                },
            )
        )

    def test_sheets_overlay_fills_blank_dob_wean_before_cage_card_generation(self) -> None:
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            config, raw_one, raw_two, fake_translation, inventory_path = (
                self._build_two_mouse_fixture(root)
            )

            # Blank out DOB/Wean_By in the local inventory copy so the only
            # source for those two fields is the (fake) Google Sheet.
            rows = inventory_path.read_text(encoding="utf-8").splitlines()
            reader = list(csv.reader(rows))
            for row in reader[1:]:
                row[15] = ""
                row[17] = ""
            with inventory_path.open("w", encoding="utf-8", newline="") as stream:
                csv.writer(stream).writerows(reader)

            credentials_file = root / "service_account.json"
            credentials_file.write_text("{}", encoding="utf-8")
            config = AppConfig(
                project_root=config.project_root,
                runtime_root=config.runtime_root,
                application_version=config.application_version,
                r=config.r,
                transnetyx=config.transnetyx,
                translation=config.translation,
                inventory=config.inventory,
                cage_card=config.cage_card,
                sheets_overlay=SheetsOverlayConfig(
                    enabled=True,
                    spreadsheet_id="abc123",
                    worksheet="Sheet1",
                    credentials_file=credentials_file,
                ),
            )

            self._install_fake_google_modules(
                sheet_rows=[
                    ["Mouse", "DOB", "Wean_By"],
                    ["CM0001", "07/01/2026", "07/29/2026"],
                    ["CM0002", "07/01/2026", "07/29/2026"],
                ]
            )

            with patch("automouse.app.run_r_translation", side_effect=fake_translation):
                context = run_batch([raw_one, raw_two], config)

            self.assertEqual(context.stage, RunStage.COMPLETED)
            self.assertFalse(
                any("Could not refresh DOB/Wean_By" in warning for warning in context.warnings)
            )
            cards = load_workbook(context.artifacts["cage_card_file"], read_only=True)
            sheet = cards["Sheet1"]
            self.assertIsNotNone(sheet.cell(2, 6).value)  # DATE WEANED
            self.assertIsNotNone(sheet.cell(2, 7).value)  # DATE BORN
            cards.close()

            # The overlay must fill the in-memory table used for cage cards
            # without corrupting the master inventory's own file on disk
            # beyond the normal, already-reviewed inventory-update write.
            updated_inventory = Path(context.artifacts["updated_inventory_csv_file"])
            updated_rows = list(csv.reader(updated_inventory.read_text(encoding="utf-8").splitlines()))
            self.assertEqual(updated_rows[1][15], "07/01/2026")
            self.assertEqual(updated_rows[1][17], "07/29/2026")

            # Per-mouse-ID audit trail for what the overlay actually filled,
            # not just an aggregate count.
            fills = context.environment["sheets_overlay_fills"]
            self.assertEqual(len(fills), 4)  # DOB + Wean_By for each of 2 mice
            self.assertTrue(any("CM0001" in message for message in fills))
            self.assertTrue(any("CM0002" in message for message in fills))

            summary = json.loads(
                Path(context.artifacts["run_summary_file"]).read_text(encoding="utf-8")
            )
            self.assertTrue(summary["sheets_overlay_enabled"])
            self.assertEqual(len(summary["sheets_overlay_fills"]), 4)
            self.assertEqual(summary["application_version"], "test")
            self.assertEqual(summary["python_version"], sys.version.split()[0])
            self.assertIn("translated_output_file", summary["output_sha256_by_file"])
            self.assertEqual(
                summary["translation_script_sha256"],
                hashlib.sha256(config.r.translation_script.read_bytes()).hexdigest(),
            )

    def test_sheets_overlay_fetch_failure_warns_and_still_completes(self) -> None:
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            config, raw_one, raw_two, fake_translation, _ = self._build_two_mouse_fixture(root)

            config = AppConfig(
                project_root=config.project_root,
                runtime_root=config.runtime_root,
                application_version=config.application_version,
                r=config.r,
                transnetyx=config.transnetyx,
                translation=config.translation,
                inventory=config.inventory,
                cage_card=config.cage_card,
                sheets_overlay=SheetsOverlayConfig(
                    enabled=True,
                    spreadsheet_id="abc123",
                    worksheet="Sheet1",
                    credentials_file=root / "does_not_exist.json",
                ),
            )

            with patch("automouse.app.run_r_translation", side_effect=fake_translation):
                context = run_batch([raw_one, raw_two], config)

            self.assertEqual(context.stage, RunStage.COMPLETED)
            self.assertTrue(
                any("Could not refresh DOB/Wean_By" in warning for warning in context.warnings)
            )


if __name__ == "__main__":
    unittest.main()
