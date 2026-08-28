from __future__ import annotations

import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from automouse.config import InventoryConfig
from automouse.inventory_manager import InventoryTable, save_updated_inventory


class InventoryManagerTests(unittest.TestCase):
    def test_saved_inventory_is_sorted_by_numeric_mouse_id(self) -> None:
        with tempfile.TemporaryDirectory() as directory_name:
            root = Path(directory_name)
            config = InventoryConfig(
                file=root / "inventory.csv",
                columns={"mouse_id": 1, "genotype": 2},
            )
            table = InventoryTable(
                root / "inventory.csv",
                ["Mouse", "Genotype"],
                [
                    ["CM10", "K/+"],
                    ["CM2", "+/+"],
                    ["CM1", "K/+"],
                ],
                config,
            )

            csv_path, xlsx_path = save_updated_inventory(
                table,
                root / "updated.csv",
                root / "updated.xlsx",
            )

            with csv_path.open(encoding="utf-8-sig", newline="") as stream:
                rows = list(csv.reader(stream))
            self.assertEqual([row[0] for row in rows[1:]], ["CM1", "CM2", "CM10"])

            workbook = load_workbook(xlsx_path, read_only=True)
            sheet = workbook["Mouse Inventory"]
            self.assertEqual(
                [sheet.cell(row, 1).value for row in range(2, 5)],
                ["CM1", "CM2", "CM10"],
            )
            workbook.close()


if __name__ == "__main__":
    unittest.main()
